import io
import os
import cv2
import numpy as np
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from PIL import Image
from imutils.perspective import four_point_transform

def get_template_path():
    base_dir = os.path.dirname(os.path.dirname(__file__))
    return os.path.join(base_dir, 'template', 'GASTOS_VIAJE.xlsx')

def obtener_celda_principal(hoja, celda):
    """Obtiene la celda principal si está en un rango fusionado"""
    for merged_range in hoja.merged_cells.ranges:
        if celda.coordinate in merged_range:
            return hoja.cell(merged_range.min_row, merged_range.min_col)
    return celda

def fill_excel_template(data):
    # Logging de datos entrantes
    print("=" * 50)
    print("DATOS RECIBIDOS EN fill_excel_template:")
    print(f"Data completa: {data}")
    print("=" * 50)
    
    wb = load_workbook(get_template_path())
    ws = wb.active

    def obtener_celda_principal(hoja, celda):
        for merged_range in hoja.merged_cells.ranges:
            if celda.coordinate in merged_range:
                return hoja.cell(merged_range.min_row, merged_range.min_col)
        return celda

    # Llenar campos básicos
    campos = {
        'E4': data.get('empresa', ''),
        'E6': data.get('nit', ''),
        'B7': data.get('placa', ''),
        'H7': data.get('conductor', ''),
        'E10': data.get('desde', ''),
        'I10': data.get('hasta', ''),
        'B13': data.get('fecha', ''),
        'G14': data.get('anticipo', ''),
        'J14': data.get('flete', '')
    }
    for celda, valor in campos.items():
        cell = ws[celda]
        main_cell = obtener_celda_principal(ws, cell)
        main_cell.value = valor

    # Llenar gastos
    gastos = data.get('gastos', {})
    print("\nGASTOS RECIBIDOS:")
    print(f"gastos: {gastos}")
    print(f"Tipo de gastos: {type(gastos)}")
    
    mapping = {
        'acpm': 'G16',
        'cargue': 'G18',
        'descargue': 'G20',
        'peajes': 'G22',
        'comision_empresa': 'G24',
        'llantas': 'G26',
        'engrase': 'G28',
        'lavada': 'G30',
        'parqueadero': 'G32',
        'carrosada': 'G34',
        'descarrrosada': 'G36',
        'otros': 'G38',
        'bonificacion': 'G40'
    }
    
    print("\nLLENANDO CELDAS DE GASTOS:")
    for key, celda in mapping.items():
        valor = gastos.get(key, 0)
        print(f"  {key} -> celda {celda} = {valor}")
        cell = ws[celda]
        main_cell = obtener_celda_principal(ws, cell)
        main_cell.value = valor
    print("=" * 50 + "\n")

    # Calcular totales
    flete = float(data.get('flete', 0) or 0)
    anticipo = float(data.get('anticipo', 0) or 0)
    bonificacion = float(gastos.get('bonificacion', 0) or 0)
    total_gastos = sum(float(gastos.get(k, 0) or 0) for k in mapping.keys())

    valor_viaje = flete + bonificacion
    menos_anticipo = anticipo - total_gastos  # Anticipo menos los gastos
    
    # Desde la perspectiva de la EMPRESA:
    # Si menos_anticipo es positivo: sobró dinero = saldo a favor (empresa debe recibir)
    # Si menos_anticipo es negativo: gastó más del anticipo = saldo en contra (empresa debe pagar)
    saldo_a_favor = menos_anticipo if menos_anticipo > 0 else 0
    saldo_en_contra = abs(menos_anticipo) if menos_anticipo < 0 else 0

    resumen = {
        'I41': valor_viaje,
        'I42': total_gastos,
        'I43': menos_anticipo,
        'I44': saldo_a_favor,
        'I45': saldo_en_contra
    }
    for celda, valor in resumen.items():
        cell = ws[celda]
        main_cell = obtener_celda_principal(ws, cell)
        main_cell.value = valor

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def process_image(img_data):
    """
    Procesa una imagen para detectar y recortar el documento
    """
    try:
        # Convertir bytes a imagen OpenCV
        pil_image = Image.open(io.BytesIO(img_data))
        image = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)
        
        # Contorno por defecto (imagen completa)
        height, width = image.shape[:2]
        document_contour = np.array([[0, 0], [width, 0], [width, height], [0, height]])
        
        # Preprocesamiento suave
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (3, 3), 0)
        
        # Detección de bordes
        edges = cv2.Canny(blur, 30, 100, apertureSize=3)
        
        # Dilatación mínima
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        edges = cv2.dilate(edges, kernel, iterations=1)
        
        # Encontrar contornos
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contours = sorted(contours, key=cv2.contourArea, reverse=True)
        
        max_area = 0
        image_area = width * height
        
        # Buscar contorno del documento
        for contour in contours[:15]:
            area = cv2.contourArea(contour)
            
            if area > image_area * 0.25:
                peri = cv2.arcLength(contour, True)
                approx = cv2.approxPolyDP(contour, 0.015 * peri, True)
                
                if len(approx) >= 4 and area > max_area:
                    if len(approx) > 4:
                        contour_points = approx.reshape(-1, 2)
                        tl = contour_points[np.argmin(contour_points[:, 0] + contour_points[:, 1])]
                        tr = contour_points[np.argmin(contour_points[:, 1] - contour_points[:, 0])]
                        br = contour_points[np.argmax(contour_points[:, 0] + contour_points[:, 1])]
                        bl = contour_points[np.argmax(contour_points[:, 1] - contour_points[:, 0])]
                        approx = np.array([tl, tr, br, bl]).reshape(4, 1, 2)
                    
                    document_contour = approx
                    max_area = area
        
        # Aplicar transformación de perspectiva
        warped = four_point_transform(image, document_contour.reshape(4, 2))
        
        # Aplicar margen mínimo
        h, w = warped.shape[:2]
        margin_percent = 0.005
        margin = max(2, int(min(h, w) * margin_percent))
        
        if h > 2*margin and w > 2*margin:
            cropped = warped[margin:h-margin, margin:w-margin]
        else:
            cropped = warped
        
        # Mejorar contraste
        gray_cropped = cv2.cvtColor(cropped, cv2.COLOR_BGR2GRAY)
        clahe = cv2.createCLAHE(clipLimit=1.5, tileGridSize=(8, 8))
        enhanced = clahe.apply(gray_cropped)
        final_image = cv2.cvtColor(enhanced, cv2.COLOR_GRAY2BGR)
        
        # Asegurar que la imagen sea vertical (retrato)
        h, w = final_image.shape[:2]
        if w > h:  # Si es horizontal, rotarla 90 grados
            final_image = cv2.rotate(final_image, cv2.ROTATE_90_CLOCKWISE)
        
        # Convertir a bytes
        is_success, buffer = cv2.imencode(".jpg", final_image)
        if is_success:
            return io.BytesIO(buffer).getvalue()
        
        return img_data  # Retornar imagen original si falla el procesamiento
        
    except Exception as e:
        print(f"Error procesando imagen: {e}")
        return img_data  # Retornar imagen original si hay error

def generate_invoice_pdf(images_data):
    """
    Genera un PDF con las imágenes de facturas organizadas dinámicamente manteniendo su proporción.
    images_data: Lista de bytes de las imágenes
    """
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Configuración de márgenes y espaciado
    margin = 20  # Reducido para más espacio
    spacing = 15
    max_width = width - (2 * margin)
    max_height = height - (2 * margin)
    
    # TAMAÑOS REDUCIDOS para que quepan más imágenes por página
    # Reducción del 40% de los tamaños anteriores
    max_image_width = max_width * 0.48   # Reducido 40% desde 0.8
    max_image_height = max_height * 0.36  # Reducido 40% desde 0.6

    current_x = margin
    current_y = height - margin
    current_row_height = 0
    page_number = 1

    for img_data in images_data:
        # Procesar la imagen para detectar y recortar el documento
        processed_img_data = process_image(img_data)
        
        # Procesar la imagen para obtener sus dimensiones originales
        img = Image.open(io.BytesIO(processed_img_data))
        original_width, original_height = img.size
        aspect_ratio = original_width / original_height

        # Calcular el tamaño final manteniendo la proporción
        if aspect_ratio > 1:  # Imagen horizontal
            final_width = min(max_image_width, max_width * 0.45)  # Reducido 40% desde 0.75
            final_height = final_width / aspect_ratio
        else:  # Imagen vertical (típico para facturas de celular)
            # Para facturas verticales, usar espacio reducido
            final_width = min(max_image_width, max_width * 0.42)   # Reducido 40% desde 0.7
            final_height = final_width / aspect_ratio
            # Si el alto se pasa, limitar el espacio vertical
            if final_height > max_height * 0.48:  # Reducido 40% desde 0.8
                final_height = max_height * 0.48
                final_width = final_height * aspect_ratio

        # Verificar si la imagen cabe en la fila actual
        if current_x + final_width > width - margin:
            current_x = margin
            current_y -= (current_row_height + spacing)
            current_row_height = 0

        # Verificar si la imagen cabe en la página actual
        if current_y - final_height < margin:
            c.showPage()
            current_x = margin
            current_y = height - margin
            current_row_height = 0
            page_number += 1

        # Dibujar la imagen
        y_position = current_y - final_height
        img_reader = ImageReader(img)
        c.drawImage(img_reader, current_x, y_position, width=final_width, height=final_height)

        # Actualizar posiciones
        current_x += final_width + spacing
        current_row_height = max(current_row_height, final_height)

    c.save()
    buffer.seek(0)
    return buffer

def generate_exportable_excel(exportable_data):
    """
    Genera un archivo Excel con los datos contables del exportable
    Formato: Cuenta | Comprobante | Fecha | Documento | DocumentoRef | Nit | Detalle | Tipo | Valor | Base | Centro de Costo | Trans. Ext | Plazo
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from datetime import datetime
    
    # Crear nuevo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = exportable_data.get('sheetName', 'Exportable Contable')
    
    # Obtener datos
    data = exportable_data.get('data', [])
    headers = exportable_data.get('headers', [
        'Cuenta', 'Comprobante', 'Fecha(mm/dd/yyyy)', 'Documento', 
        'Documento Ref', 'Nit', 'Detalle', 'Tipo', 'Valor', 'Base', 
        'Centro de Costo', 'Trans. Ext', 'Plazo'
    ])
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Escribir datos
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            # Mapear el header al campo correspondiente
            field_mapping = {
                'Cuenta': 'Cuenta',
                'Comprobante': 'Comprobante',
                'Fecha(mm/dd/yyyy)': 'Fecha(mm/dd/yyyy)',
                'Documento': 'Documento',
                'Documento Ref': 'Documento Ref',
                'Nit': 'Nit',
                'Detalle': 'Detalle',
                'Tipo': 'Tipo',
                'Valor': 'Valor',
                'Base': 'Base',
                'Centro de Costo': 'Centro de Costo',
                'Trans. Ext': 'Trans. Ext',
                'Plazo': 'Plazo'
            }
            
            field_name = field_mapping.get(header, header)
            value = row_data.get(field_name, '')
            
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border_style
            
            # Formatear números
            if header in ['Valor', 'Base'] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            elif header == 'Tipo' and isinstance(value, (int, float)):
                cell.number_format = '0'
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 12,  # Cuenta
        'B': 12,  # Comprobante
        'C': 15,  # Fecha
        'D': 12,  # Documento
        'E': 12,  # Documento Ref
        'F': 15,  # Nit
        'G': 30,  # Detalle
        'H': 8,   # Tipo
        'I': 12,  # Valor
        'J': 12,  # Base
        'K': 15,  # Centro de Costo
        'L': 12,  # Trans. Ext
        'M': 8    # Plazo
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer